# The following functions are used for common tasks with respect to creating and evaluating the
# Excel files for the PC Practical Course
#
# get_Matrkelnummern():
#       opens a file with a list of all given student-IDs (Matrikelnummern)
#
# create_workbooks_in_Abgaben():
#       creates a directory called "Abgaben" and subdirectories called according to the student-IDs.
#       In each of these subdirectories an Excel-File named matrnr.xlsx will be placed. This structure
#       is for testing purposes as the LMS Ilias delivers the results by the students accordingly.
#       The checking of the Excel-Files is based on this directory-structure.
#
# create_workbooks_in_Excel_Files():
#       Similar to the above function. Just the directory structure is different and better usable
#       for distributing the Excel-Files to the students.
#
# check_Excel_Files_in_Abgaben():
#       Uses the directory structure from above. Goes through this structure and opens all Excel-Files.
#
# check_excel_file():
#       Actually opens the excel file and executes the checking-functions defined by the parameter func.
#       This function must be implemented for each exercise
#
# result_into_file():
#       The awarded points per exercise are handed over in a list. This list will be transformed into an 
#       Excel-File for further manual analysis.


from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
import os

def get_Matrikelnummern(filename):
    # Datei öffnen und eine Liste dieser Matrikelnummern erzeugen
    l_matr = []
    with open(filename, "r", encoding="utf-8-sig") as file:
        for line in file:
            line = line.strip()
            matr = int(line)
            l_matr.append(matr)
            l_matr.sort()
    return l_matr

def create_work_books_in_Abgaben(l_matr, func):
    # Create a directory called "Abgaben", in this directory subdirectories called matr and in these subdirs excel-files called matr.xlsx
    if os.path.exists("Abgaben"):
        print("Directory named 'Abgaben' already exists, no files created.")
    else:
        print("Go for it")
        os.makedirs("Abgaben")
        os.chdir("Abgaben")
        for matr in l_matr:
            matr = str(matr)
            os.makedirs(matr)
            os.chdir(matr)
            filename = matr + ".xlsx"
            wb = func(filename, matr)
            wb.save(filename)
            os.chdir("..")
        os.chdir("..")
        print("Excel-Files created")

def create_work_books_in_Excel_Files(l_matr, func):
    # Create a directory called "Excel_Files", in this directory create Excel-Files called matr.xlsx
    if os.path.exists("Excel_Files"):
        print("Directory named 'Excel-Files' already exists, no files created.")
    else:
        print("Go for it")
        os.makedirs("Excel_Files")
        os.chdir("Excel_Files")
        for matr in l_matr:
            matr = str(matr)
            filename = matr + ".xlsx"
            wb = func(filename, matr)
            wb.save(filename)
        os.chdir("..")
        print("Excel-Files created")


def check_excel_file(filename, func):
    if filename[7:] == ".xlsx":
        try:
            matrnr = int(filename[:7])
            wb = load_workbook(filename)
            wb_val = load_workbook(filename, data_only = True)
            points = func(matrnr, wb, wb_val)
            return points
        except:
            print("Problems with ", filename, "! Cannot be opened by openpyxl")
    return 0    

def check_ExcelFiles_in_Abgaben(func):
    l_points = []
    if os.path.exists("Abgaben"):
        os.chdir("Abgaben")
        for subdir in os.listdir():
            if subdir != ".DS_Store":
                os.chdir(subdir)
                for filename in os.listdir():
                    if filename != ".DS_Store":
                        points = check_excel_file(filename, func)
                        l_points.append(points)
                os.chdir("..")
        os.chdir("..")
    else:
        print("No directory named 'Abgaben'.")
    return l_points

def result_into_file(l_points, header):
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for points in l_points:
        ws.append(points)
    
    # Some formatting
    highlight = NamedStyle(name = "highlight")
    highlight.font = Font(bold = True, color = "FFFFFF")
    highlight.fill = PatternFill(fgColor = "00b1ac", fill_type = "solid")
    highlight.alignment = Alignment(horizontal = "center")
    wb.add_named_style(highlight)
    
    for col in ws.iter_cols(min_row = 1, max_row = 1, max_col = len(header)):
        for cell in col:
            cell.style = "highlight"
            
    wb.save("Results.xlsx")
